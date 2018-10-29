from openpyxl import load_workbook
import os
import datetime
from datetime import date

today = datetime.date.today()
today_date = int(today.strftime("%d"))

#The function to work with a relative paths
THIS_FOLDER = os.path.dirname(os.path.relpath(__file__))
grafik = os.path.join(THIS_FOLDER, 'python-report-1.xlsx')

# Load in the workbook
wb = load_workbook (grafik, data_only=True)

#sheet name
ws = wb['ИЮЛЬ']

#the values for defining a cashier
cash_position1 = 'Контролёр-кассир 4-го разряда'
cash_position2 = 'Старшй кассир 5-го разряда'


#reading values
for i in range(14, ws.max_row):
  cell_number = ws.cell(row=i, column=1).value
  cell_next_number = ws.cell(row=(i+1), column=1).value
  cell_position = ws.cell(row=i, column=4).value
  cell_last_name =  ws.cell(row=i, column=2).value
  
  #print out all the cashiers + working hours
  if cell_number is not None and cell_position == cash_position1 or cell_position == cash_position2:
    for h in range (6, ws.max_column):
      cell_calendar_date = ws.cell(row=12, column=h).value
      cell_working_hours_today = ws.cell(row=i, column=h).value
      if cell_calendar_date == today_date:
        print (cell_last_name)
        print (cell_working_hours_today)

  #breaking the cycle in the end of the positions
  if cell_number is None and cell_next_number is None:
    break